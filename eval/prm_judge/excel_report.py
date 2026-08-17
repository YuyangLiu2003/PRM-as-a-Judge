"""生成 PRM-as-a-Judge 标准指标工作簿。

用法示例：
    write_metrics_workbook(run_root / "metrics.xlsx", records, aggregate, metric_config)
"""

from __future__ import annotations

import math
import os
import tempfile
from collections import defaultdict
from copy import copy
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from .metrics import (
    MODEL_SUMMARY_METRICS,
    MetricConfig,
    aggregate_median_model_metrics,
    metric_value,
    record_has_drawdown,
    trace_summary,
)

CASE_METRICS = list(MODEL_SUMMARY_METRICS)
SUMMARY_METRICS = list(MODEL_SUMMARY_METRICS)
CASE_COLUMNS = [
    "benchmark",
    "model",
    "task_name",
    "case_id",
    "task",
    "label",
    "status",
    "error",
    "num_points",
    *CASE_METRICS,
]
TASK_SUMMARY_COLUMNS = [
    "benchmark",
    "model",
    "task",
    *SUMMARY_METRICS,
]
MODEL_SUMMARY_COLUMNS = [
    "benchmark",
    "model",
    *SUMMARY_METRICS,
]
LOWER_IS_BETTER_METRICS = {"CRA", "STR"}

HEADER_FILL = PatternFill("solid", fgColor="335C9F")
HEADER_FONT = Font(name="Aptos", size=10, bold=True, color="FFFFFF")
BODY_FONT = Font(name="Aptos", size=10, color="1F2937")
BANDED_FILL = PatternFill("solid", fgColor="EAF2FB")
ERROR_FILL = PatternFill("solid", fgColor="FDECEC")
PERCENT_FORMAT = "0.00%"
COUNT_FORMAT = "#,##0"


def _group_key(record: dict[str, Any]) -> tuple[str, str, str]:
    """返回与指标聚合函数一致的 benchmark/model/task 分组键。"""
    return (
        str(record.get("benchmark", "manifest")),
        str(record.get("model", "default")),
        str(record.get("task_name") or record.get("case_task") or "all"),
    )


def _model_key(record: dict[str, Any]) -> tuple[str, str]:
    """返回与 leaderboard 一致的 benchmark/model 分组键。"""
    benchmark, model, _ = _group_key(record)
    return benchmark, model


def _finite_number(value: Any) -> float | int | None:
    """把有限数值保留为数值，其余内容转换为空白。"""
    if isinstance(value, bool):
        return int(value)
    if not isinstance(value, (int, float)):
        return None
    numeric = float(value)
    if not math.isfinite(numeric):
        return None
    return value


def _sr_sort_key(row: dict[str, Any]) -> tuple[bool, float, str, str, str]:
    """返回按 SR 降序、缺失值最后的稳定排序键。"""
    sr = _finite_number(row.get("SR"))
    return (
        sr is None,
        -float(sr) if sr is not None else 0.0,
        str(row.get("benchmark", "")),
        str(row.get("model", "")),
        str(row.get("task", "")),
    )


def _is_lower_better_metric(metric: str) -> bool:
    """判断模型级展示列是否以较小值为优。"""
    return metric in LOWER_IS_BETTER_METRICS


def _case_row(record: dict[str, Any], config: MetricConfig) -> dict[str, Any]:
    """构造一个逐 case 行，错误 case 的指标保持为空。"""
    progress = record.get("progress_processed")
    if not isinstance(progress, list):
        progress = record.get("progress_raw")
    row: dict[str, Any] = {
        "benchmark": str(record.get("benchmark", "manifest")),
        "model": str(record.get("model", "default")),
        "task_name": str(record.get("task_name") or record.get("case_task") or "all"),
        "case_id": str(record.get("case_id", "")),
        "task": str(record.get("task") or record.get("case_task") or ""),
        "label": record.get("label"),
        "status": str(record.get("status") or "error"),
        "error": record.get("error"),
        "num_points": len(progress) if isinstance(progress, list) else 0,
    }
    if record.get("status") != "ok" or not isinstance(record.get("metrics"), dict):
        row.update({metric: None for metric in CASE_METRICS})
        return row

    metrics = record["metrics"]
    diagnostics = trace_summary(record, config)
    row.update(
        {
            metric: _finite_number(metric_value(metrics, metric))
            for metric in CASE_METRICS
        }
    )
    row["SR"] = _finite_number(diagnostics["SR_trace"])
    row["DRR"] = (
        _finite_number(diagnostics["DRR_trace"])
        if record_has_drawdown(record, config)
        else None
    )
    row["FNS"] = _finite_number(diagnostics["FNS_trace"])
    row["SQS"] = _finite_number(diagnostics["SQS_trace"])
    return row


def build_workbook_rows(
    records: list[dict[str, Any]],
    aggregate: dict[str, Any],
    config: MetricConfig,
) -> dict[str, list[dict[str, Any]]]:
    """构造逐 case 与任务汇总行，并保留只含错误 case 的分组。"""
    case_rows = [_case_row(record, config) for record in records]
    task_keys = {_group_key(record) for record in records}

    task_metrics = {
        (str(row["benchmark"]), str(row["model"]), str(row["task"])): row
        for row in aggregate.get("groups", [])
    }
    task_rows: list[dict[str, Any]] = []
    for benchmark, model, task in sorted(task_keys):
        metric_row = task_metrics.get((benchmark, model, task))
        row: dict[str, Any] = {
            "benchmark": benchmark,
            "model": model,
            "task": task,
        }
        row.update(
            {
                metric: _finite_number(metric_row.get(metric)) if metric_row is not None else None
                for metric in SUMMARY_METRICS
            }
        )
        task_rows.append(row)

    task_rows.sort(key=_sr_sort_key)
    return {
        "Cases": case_rows,
        "Task Summary": task_rows,
    }


def build_model_summary_rows(
    records: list[dict[str, Any]],
    leaderboard: list[dict[str, Any]],
    metric_names: list[str] | tuple[str, ...],
) -> list[dict[str, Any]]:
    """按给定模型级指标列构造行，并保留只含错误 case 的模型。"""
    model_keys = {_model_key(record) for record in records}
    model_metrics = {
        (str(row["benchmark"]), str(row["model"])): row
        for row in leaderboard
    }

    rows: list[dict[str, Any]] = []
    for benchmark, model in sorted(model_keys):
        metric_row = model_metrics.get((benchmark, model))
        row: dict[str, Any] = {
            "benchmark": benchmark,
            "model": model,
        }
        row.update(
            {
                metric: _finite_number(metric_row.get(metric)) if metric_row is not None else None
                for metric in metric_names
            }
        )
        rows.append(row)
    rows.sort(key=_sr_sort_key)
    return rows


def _write_text_cell(cell: Any, value: Any) -> None:
    """显式按纯文本写入用户字段，避免以公式控制字符开头时被执行。"""
    if value is None:
        cell.value = None
        return
    cell.value = str(value)
    cell.data_type = "s"


def _apply_table_style(
    worksheet: Any,
    columns: list[str],
    rows: list[dict[str, Any]],
    *,
    table_name: str,
    freeze_panes: str,
    widths: dict[str, float],
    color_scale_metric: str | None = "SR",
    rank_metrics: bool = False,
) -> None:
    """写入平面数据表并应用统一的筛选、冻结、配色和数字格式。"""
    worksheet.sheet_view.showGridLines = False
    worksheet.freeze_panes = freeze_panes
    worksheet.sheet_view.zoomScale = 90
    worksheet.append(columns)

    metric_columns = set(columns) & (set(CASE_METRICS) | set(SUMMARY_METRICS))
    text_columns = set(columns) - metric_columns - {"num_points"}
    count_columns = {"num_points"}
    for row_index, row in enumerate(rows, start=2):
        for column_index, column in enumerate(columns, start=1):
            cell = worksheet.cell(row=row_index, column=column_index)
            value = row.get(column)
            if column in text_columns:
                _write_text_cell(cell, value)
            else:
                cell.value = _finite_number(value)
            cell.font = BODY_FONT
            cell.alignment = Alignment(
                horizontal="left" if column in text_columns else "right",
                vertical="top",
                wrap_text=column in {"task", "error"},
            )
            if column in metric_columns:
                cell.number_format = PERCENT_FORMAT
            elif column in count_columns:
                cell.number_format = COUNT_FORMAT
        if row_index % 2 == 0:
            for cell in worksheet[row_index]:
                cell.fill = BANDED_FILL
        if row.get("status") != "ok" and "status" in columns:
            worksheet.cell(row=row_index, column=columns.index("status") + 1).fill = ERROR_FILL
            worksheet.cell(row=row_index, column=columns.index("error") + 1).fill = ERROR_FILL
            worksheet.row_dimensions[row_index].height = 42

    for cell in worksheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    worksheet.row_dimensions[1].height = 30

    for index, column in enumerate(columns, start=1):
        worksheet.column_dimensions[get_column_letter(index)].width = widths.get(column, 12)

    last_row = len(rows) + 1
    last_column = get_column_letter(len(columns))
    table_ref = f"A1:{last_column}{last_row}"
    worksheet.auto_filter.ref = table_ref
    if rows:
        table = Table(displayName=table_name, ref=table_ref)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        worksheet.add_table(table)

    if rank_metrics and rows:
        _apply_metric_rank_styles(worksheet, columns, rows, metric_columns)

    if rows and color_scale_metric is not None and color_scale_metric in columns:
        metric_column = get_column_letter(columns.index(color_scale_metric) + 1)
        worksheet.conditional_formatting.add(
            f"{metric_column}2:{metric_column}{last_row}",
            ColorScaleRule(
                start_type="num",
                start_value=0,
                start_color="FFFFFF",
                mid_type="num",
                mid_value=0.5,
                mid_color="D9EAF7",
                end_type="num",
                end_value=1,
                end_color="5B9BD5",
            ),
        )


def _apply_metric_rank_styles(
    worksheet: Any,
    columns: list[str],
    rows: list[dict[str, Any]],
    metric_columns: set[str],
) -> None:
    """按 benchmark 标记各指标最优值为粗体、次优值为下划线。"""
    grouped_rows: dict[str, list[tuple[int, dict[str, Any]]]] = defaultdict(list)
    for row_index, row in enumerate(rows, start=2):
        grouped_rows[str(row.get("benchmark", ""))].append((row_index, row))

    for items in grouped_rows.values():
        for metric in metric_columns:
            if metric not in columns:
                continue
            values = {
                float(value)
                for _, row in items
                if (value := _finite_number(row.get(metric))) is not None
            }
            if not values:
                continue
            ordered = sorted(values, reverse=not _is_lower_better_metric(metric))
            best = ordered[0]
            second = ordered[1] if len(ordered) > 1 else None
            column_index = columns.index(metric) + 1
            for row_index, row in items:
                value = _finite_number(row.get(metric))
                if value is None:
                    continue
                cell = worksheet.cell(row=row_index, column=column_index)
                if float(value) == best:
                    ranked_font = copy(cell.font)
                    ranked_font.bold = True
                    cell.font = ranked_font
                elif second is not None and float(value) == second:
                    ranked_font = copy(cell.font)
                    ranked_font.underline = "single"
                    cell.font = ranked_font


def _save_workbook_atomic(workbook: Workbook, path: Path) -> None:
    """将工作簿写入同目录临时文件后原子替换目标文件。"""
    path.parent.mkdir(parents=True, exist_ok=True)
    temp_path: Path | None = None
    try:
        with tempfile.NamedTemporaryFile(
            prefix=f".{path.stem}.",
            suffix=".tmp.xlsx",
            dir=path.parent,
            delete=False,
        ) as file_obj:
            temp_path = Path(file_obj.name)
        workbook.save(temp_path)
        os.replace(temp_path, path)
    finally:
        workbook.close()
        if temp_path is not None and temp_path.exists():
            temp_path.unlink()


def write_metrics_workbook(
    path: Path,
    records: list[dict[str, Any]],
    aggregate: dict[str, Any],
    config: MetricConfig | None = None,
) -> None:
    """原子写出 Cases、Task Summary、模型均值与模型中位数工作表。"""
    cfg = config or MetricConfig()
    workbook_rows = build_workbook_rows(records, aggregate, cfg)
    mean_rows = build_model_summary_rows(
        records,
        aggregate.get("leaderboard", []),
        SUMMARY_METRICS,
    )
    median_aggregate = aggregate_median_model_metrics(records, cfg)
    median_rows = build_model_summary_rows(
        records,
        median_aggregate.get("leaderboard", []),
        SUMMARY_METRICS,
    )
    workbook = Workbook()
    workbook.remove(workbook.active)

    cases = workbook.create_sheet("Cases")
    _apply_table_style(
        cases,
        CASE_COLUMNS,
        workbook_rows["Cases"],
        table_name="CasesTable",
        freeze_panes="E2",
        widths={
            "benchmark": 18,
            "model": 24,
            "task_name": 30,
            "case_id": 52,
            "task": 48,
            "label": 14,
            "status": 12,
            "error": 48,
            "num_points": 12,
        },
    )

    task_summary = workbook.create_sheet("Task Summary")
    _apply_table_style(
        task_summary,
        TASK_SUMMARY_COLUMNS,
        workbook_rows["Task Summary"],
        table_name="TaskSummaryTable",
        freeze_panes="D2",
        widths={"benchmark": 18, "model": 24, "task": 30},
        color_scale_metric="SR",
    )

    model_summary = workbook.create_sheet("Model Summary Mean")
    _apply_table_style(
        model_summary,
        MODEL_SUMMARY_COLUMNS,
        mean_rows,
        table_name="ModelSummaryMeanTable",
        freeze_panes="C2",
        widths={"benchmark": 18, "model": 24},
        color_scale_metric="SR",
        rank_metrics=True,
    )

    model_summary_median = workbook.create_sheet("Model Summary Median")
    _apply_table_style(
        model_summary_median,
        MODEL_SUMMARY_COLUMNS,
        median_rows,
        table_name="ModelSummaryMedianTable",
        freeze_panes="C2",
        widths={"benchmark": 18, "model": 24},
        color_scale_metric="SR",
        rank_metrics=True,
    )

    _save_workbook_atomic(workbook, path)
